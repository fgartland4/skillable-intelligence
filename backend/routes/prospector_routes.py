"""Prospector blueprint — batch company scoring and results.

Routes are thin surfaces: parse input → call Intelligence → render template.
All research, scoring, caching, and academic logic lives in intelligence.py.
"""

import csv
import io
import threading
import uuid

from flask import Blueprint, render_template, request, redirect, url_for, Response, stream_with_context, jsonify

from core import _push, _sse_stream, _attach_scores, _compute_composite, _fmt_ondemand, _fmt_cert, _cancelled_jobs
from intelligence import qualify as intel_qualify
from storage import (
    save_analysis, load_analysis,
    find_analysis_by_company_name,
    save_prospector_run, load_prospector_run,
    append_poor_fit_feedback, load_poor_fit_companies,
    save_discovery, load_discovery,
    load_competitor_candidates, clear_competitor_candidates,
)

import datetime
import logging
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Prospector Blueprint
# ---------------------------------------------------------------------------

prospector = Blueprint("prospector", __name__, url_prefix="/prospector")


def _derive_skillable_path(lab_score: int) -> str:
    """Map a lab score to the three Prospector path labels (software companies)."""
    if lab_score >= 50:
        return "Labable"
    if lab_score >= 20:
        return "Simulations"
    return "Do Not Pursue"


# Path label formatting, academic logic, and cache date formatting
# all live in intelligence.py — imported via intel_qualify()


@prospector.route("/")
@prospector.route("")
def prospector_home():
    poor_fit = load_poor_fit_companies()
    competitor_candidates = load_competitor_candidates()
    return render_template("prospector.html",
                           poor_fit_count=len(poor_fit),
                           competitor_candidate_count=len(competitor_candidates))


@prospector.route("/run", methods=["POST"])
def prospector_run():
    company_names = []

    # Mode 1: CSV file upload — read first column
    uploaded = request.files.get("csv_file")
    if uploaded and uploaded.filename:
        stream = io.StringIO(uploaded.stream.read().decode("utf-8-sig"))
        reader = csv.reader(stream)
        for row in reader:
            if row and row[0].strip() and row[0].strip().lower() != "company":
                company_names.append(row[0].strip())

    # Mode 2: pasted text (fallback)
    if not company_names:
        raw = request.form.get("companies", "")
        company_names = [n.strip() for n in raw.replace(",", "\n").splitlines() if n.strip()]

    if not company_names:
        return redirect(url_for("prospector.prospector_home"))

    # Filter out previously flagged poor fits
    poor_fit = load_poor_fit_companies()
    filtered = [n for n in company_names if n.lower() not in poor_fit]
    skipped = len(company_names) - len(filtered)
    company_names = filtered

    if not company_names:
        return redirect(url_for("prospector.prospector_home"))

    force_refresh = request.form.get("force_refresh") == "1"

    COMPANY_TIMEOUT_SECS = 180  # 3 minutes — slow orgs are skipped, not waited on

    job_id = str(uuid.uuid4())[:8]
    results = {}
    semaphore = threading.Semaphore(6)

    def run_batch():
        def analyze_one(name):
            if job_id in _cancelled_jobs:
                return
            with semaphore:
                if job_id in _cancelled_jobs:
                    return
                _push(job_id, f"status:Analyzing {name}...")

                result_holder = [None]
                def _run():
                    result_holder[0] = intel_qualify(name, force_refresh=force_refresh)

                worker = threading.Thread(target=_run, daemon=True)
                worker.start()
                worker.join(timeout=COMPANY_TIMEOUT_SECS)

                if worker.is_alive():
                    # Timed out — release the slot and move on
                    results[name] = {"company_name": name, "timed_out": True}
                    _push(job_id, f"timeout:{name}")
                else:
                    row = result_holder[0]
                    if row:
                        # Academic path labels are set inside intelligence.qualify() — don't override
                        if not row.get("skillable_path"):
                            row["skillable_path"] = _derive_skillable_path(row.get("lab_score", 0))
                        row["flagged_poor_fit"] = False
                    results[name] = row or {"company_name": name, "error": "Analysis failed"}
                    if row and row.get("_from_cache"):
                        cache_date = row.get("_cache_date", "")
                        _push(job_id, f"cache:{name}|{cache_date}")
                    else:
                        _push(job_id, f"progress:{name}")

        threads = [threading.Thread(target=analyze_one, args=(n,), daemon=True) for n in company_names]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        cancelled = job_id in _cancelled_jobs
        _cancelled_jobs.discard(job_id)
        save_prospector_run(job_id, {
            "job_id": job_id,
            "companies": company_names,
            "skipped_poor_fit": skipped,
            "results": results,
            "status": "cancelled" if cancelled else "done",
        })
        _push(job_id, f"done:{job_id}")

    threading.Thread(target=run_batch, daemon=True).start()
    return render_template("prospector_running.html", job_id=job_id,
                           company_names=company_names,
                           company_count=len(company_names))


@prospector.route("/cancel/<job_id>", methods=["POST"])
def prospector_cancel(job_id: str):
    _cancelled_jobs.add(job_id)
    _push(job_id, f"done:{job_id}")
    return jsonify({"cancelled": True})


@prospector.route("/status/<job_id>")
def prospector_status(job_id: str):
    return Response(stream_with_context(_sse_stream(job_id, poll_interval=0.5)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@prospector.route("/results/<job_id>")
def prospector_results(job_id: str):
    job = load_prospector_run(job_id)
    if not job:
        return redirect(url_for("prospector.prospector_home"))
    results = job.get("results", {})
    rows = [r for r in results.values() if r and "error" not in r and not r.get("timed_out")]
    rows.sort(key=lambda r: r.get("composite_score", 0), reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    errors = [r for r in results.values() if r and "error" in r]
    timed_out = [r for r in results.values() if r and r.get("timed_out")]
    return render_template("prospector_results.html", rows=rows, errors=errors,
                           timed_out=timed_out,
                           job_id=job_id,
                           company_count=len(job.get("companies", [])),
                           skipped_poor_fit=job.get("skipped_poor_fit", 0))


@prospector.route("/flag", methods=["POST"])
def prospector_flag():
    data = request.get_json() or {}
    company_name = data.get("company_name", "").strip()
    job_id = data.get("job_id", "").strip()
    reason = data.get("reason", "").strip()
    if not company_name:
        return jsonify({"error": "company_name required"}), 400

    from datetime import datetime, timezone
    append_poor_fit_feedback({
        "company_name": company_name,
        "job_id": job_id,
        "reason": reason,
        "flagged_at": datetime.now(timezone.utc).isoformat(),
    })

    # Mark row as flagged in the stored job
    job = load_prospector_run(job_id)
    if job and company_name in job.get("results", {}):
        job["results"][company_name]["flagged_poor_fit"] = True
        save_prospector_run(job_id, job)

    return jsonify({"success": True})


@prospector.route("/export-csv/<job_id>")
def prospector_export_csv(job_id: str):
    job = load_prospector_run(job_id)
    if not job:
        return "Job not found", 404

    results = job.get("results", {})
    rows = sorted([r for r in results.values() if r and "error" not in r],
                  key=lambda r: r.get("composite_score", 0), reverse=True)

    output = io.StringIO()
    writer = csv.writer(output)
    # Column order mirrors the web view; exports include extra columns (titles, website)
    writer.writerow([
        "#", "Company", "Website",
        "Highly Labable", "Likely Labable", "Not Labable",
        "Top Product", "Lab Score", "Method", "Lab Maturity", "Composite",
        "ATP Program", "Channel Program", "Lab Program",
        "ILT / vILT", "On-Demand", "Gray Market", "Certs",
        "Contact 1", "Contact 1 Title", "Contact 1 LinkedIn",
        "Contact 2", "Contact 2 Title", "Contact 2 LinkedIn",
        "Score Report URL", "Notes",
    ])
    for i, row in enumerate(rows, 1):
        aid = row.get("analysis_id", "")
        writer.writerow([
            i,
            row.get("company_name", ""),
            row.get("company_url", ""),
            row.get("total_highly_labable", ""),
            row.get("total_likely_labable", ""),
            row.get("total_not_labable", ""),
            row.get("top_product", ""),
            row.get("lab_score", ""),
            row.get("skillable_path", ""),
            row.get("lab_maturity_score", ""),
            row.get("composite_score", ""),
            row.get("atp_program", ""),       # full name, not ✓
            row.get("channel_program", ""),   # full name, not ✓
            row.get("existing_lab_partner", ""),
            row.get("ilt_vilt", ""),
            row.get("ondemand_library", ""),
            row.get("gray_market", ""),
            row.get("cert_program", ""),
            row.get("top_contact_name", ""),
            row.get("top_contact_title", ""),
            row.get("top_contact_linkedin", ""),
            row.get("second_contact_name", ""),
            row.get("second_contact_title", ""),
            row.get("second_contact_linkedin", ""),
            f"/inspector/results/{aid}" if aid else "",
            "",  # Notes — blank for user
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    export_date = datetime.date.today().isoformat()
    export_filename = f"Prospector-{export_date}-{len(rows)}-{job_id[:6]}.csv"
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={export_filename}"},
    )


@prospector.route("/export/<job_id>")
def prospector_export(job_id: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    job = load_prospector_run(job_id)
    if not job:
        return "Job not found", 404

    results = job.get("results", {})
    rows = sorted([r for r in results.values() if r and "error" not in r],
                  key=lambda r: r.get("composite_score", 0), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospector Results"

    # Color palette
    green_dark  = "FF0D1A14"
    green_hi    = "FF24ED9B"
    green_mid   = "FF6b9e88"
    amber       = "FFF59E0B"
    red_soft    = "FFF87171"
    white       = "FFE8F5F0"
    muted       = "FF3d6655"
    bg_header   = "FF06100C"

    header_font  = Font(bold=True, color=green_hi, size=9)
    default_font = Font(color=white, size=9)
    muted_font   = Font(color=muted, size=8)
    thin_side    = Side(style="thin", color="1E3329")
    thin_border  = Border(bottom=thin_side)

    # Column order mirrors the web view; exports include extra columns (titles, website)
    # Col:  1    2         3        4               5               6
    headers = ["#", "Company", "Website", "Highly Labable", "Likely Labable", "Not Labable",
    #          7              8            9         10              11
               "Top Product", "Lab Score", "Method", "Lab Maturity", "Composite",
    #          12             13               14             15           16
               "ATP Program", "Channel Program", "Lab Program", "ILT / vILT", "On-Demand",
    #          17              18      19           20                  21
               "Gray Market", "Certs", "Contact 1", "Contact 1 Title", "Contact 1 LinkedIn",
    #          22           23                  24                    25              26
               "Contact 2", "Contact 2 Title", "Contact 2 LinkedIn", "Score Report", "Notes"]
    col_widths = [4, 26, 28, 8, 8, 8, 26, 10, 16, 12, 10, 30, 30, 22, 10, 12, 10, 8, 22, 26, 14, 20, 24, 14, 14, 30]

    # Centered columns: # score/count/value columns
    center_cols_hdr = {1, 4, 5, 6, 8, 9, 10, 11, 15, 16, 17, 18}

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=bg_header)
        cell.alignment = Alignment(horizontal="center" if col in center_cols_hdr else "left",
                                   vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    path_colors = {"Cloud Slice": green_hi, "VM Lab": green_hi, "Custom API": amber,
                   "Simulation": amber, "Not a Fit": red_soft}

    for i, row in enumerate(rows, 2):
        lab   = row.get("lab_score", 0)
        prtn  = row.get("lab_maturity_score", 0)
        comp  = row.get("composite_score", 0)
        path  = row.get("skillable_path", "")
        aid   = row.get("analysis_id", "")

        def score_color(s):
            return green_hi if s >= 70 else (amber if s >= 40 else red_soft)

        values = [
            row.get("rank", i - 1),           # 1  #
            row.get("company_name", ""),       # 2  Company
            row.get("company_url", ""),        # 3  Website
            row.get("total_highly_labable", ""), # 4
            row.get("total_likely_labable", ""), # 5
            row.get("total_not_labable", ""),  # 6
            row.get("top_product", ""),        # 7  Top Product
            lab,                               # 8  Lab Score
            path,                              # 9  Method
            prtn,                              # 10 Lab Maturity
            comp,                              # 11 Composite
            row.get("atp_program", ""),        # 12 ATP — full name
            row.get("channel_program", ""),    # 13 Channel — full name
            row.get("existing_lab_partner", ""), # 14 Lab Program
            row.get("ilt_vilt", ""),           # 15 ILT/vILT
            row.get("ondemand_library", ""),   # 16 On-Demand
            row.get("gray_market", ""),        # 17 Gray Market
            row.get("cert_program", ""),       # 18 Certs
            row.get("top_contact_name", ""),   # 19 Contact 1
            row.get("top_contact_title", ""),  # 20 Contact 1 Title
            row.get("top_contact_linkedin", ""), # 21 Contact 1 LinkedIn
            row.get("second_contact_name", ""), # 22 Contact 2
            row.get("second_contact_title", ""), # 23 Contact 2 Title
            row.get("second_contact_linkedin", ""), # 24 Contact 2 LinkedIn
            f"/inspector/results/{aid}" if aid else "", # 25 Score Report
            "",                                # 26 Notes
        ]
        center_cols = {1, 4, 5, 6, 8, 9, 10, 11, 15, 16, 17, 18}
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor="FF0D1A14")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col in center_cols else "left")
            if col in (8, 10, 11):
                cell.font = Font(bold=True, color=score_color(val if isinstance(val, int) else 0), size=9)
            elif col == 9:
                cell.font = Font(bold=True, color=path_colors.get(path, white), size=9)
            elif col == 1:
                cell.font = Font(color=muted, size=9, bold=True)
            elif col in (4, 5, 6):
                cell.font = Font(color=green_mid, size=9)
            else:
                cell.font = default_font
        ws.row_dimensions[i].height = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    export_date = datetime.date.today().isoformat()
    export_filename = f"Prospector-{export_date}-{len(rows)}-{job_id[:6]}.xlsx"
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={export_filename}"},
    )


# _derive_skillable_path defined at top of file — no duplicate needed here


@prospector.route("/competitor-candidates")
def competitor_candidates_json():
    """Return JSON list of competitor candidate names for pre-filling the Prospector form."""
    candidates = load_competitor_candidates()
    names = [c.get("company_name", "") for c in candidates if c.get("company_name")]
    return jsonify(names)


@prospector.route("/competitor-candidates/clear", methods=["POST"])
def competitor_candidates_clear():
    """Clear the competitor candidates list after they've been added to a run."""
    try:
        clear_competitor_candidates()
        return jsonify({"ok": True})
    except Exception as e:
        log.warning("competitor_candidates_clear failed: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500
