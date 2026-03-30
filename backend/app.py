"""Skillable Intelligence — unified Flask backend for Inspector, Designer, Prospector."""

import csv
import io
import os
import re as _re
import time
import traceback
import uuid
import threading
from pathlib import Path
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env", override=True)

import jinja2
from flask import Flask, Blueprint, render_template, request, jsonify, redirect, url_for, Response, stream_with_context
from markupsafe import Markup, escape as _escape
from researcher import discover_products, research_products, resolve_company_from_product
from scorer import discover_products_with_claude, score_selected_products
from storage import (
    save_analysis, load_analysis, list_analyses,
    save_discovery, load_discovery,
    save_batch_job, load_batch_job,
    find_analysis_by_discovery_id,
    save_prospector_run, load_prospector_run,
    append_poor_fit_feedback, load_poor_fit_companies,
)
from models import compute_labability_total

# ---------------------------------------------------------------------------
# App setup — multi-tool template loader
# ---------------------------------------------------------------------------

_BACKEND_DIR = Path(__file__).parent
_TOOLS_DIR = _BACKEND_DIR.parent / "tools"
_STATIC_DIR = _BACKEND_DIR.parent / "static"

app = Flask(
    __name__,
    static_folder=str(_STATIC_DIR),
    static_url_path="/static",
    template_folder=str(_TOOLS_DIR / "inspector" / "templates"),  # default fallback
)

# Allow templates to be found across all three tool directories
app.jinja_loader = jinja2.ChoiceLoader([
    jinja2.FileSystemLoader(str(_TOOLS_DIR / "inspector" / "templates")),
    jinja2.FileSystemLoader(str(_TOOLS_DIR / "designer" / "templates")),
    jinja2.FileSystemLoader(str(_TOOLS_DIR / "prospector" / "templates")),
])

# Custom Jinja2 tests not available in this version
app.jinja_env.tests['match'] = lambda value, pattern: bool(_re.match(pattern, str(value or '')))

# SSE stream timeout — abandon after this many seconds with no terminal message
_SSE_TIMEOUT = 600  # 10 minutes


# ---------------------------------------------------------------------------
# Template filters (registered on app, available in all blueprints)
# ---------------------------------------------------------------------------

_WARNING_LABELS = {'Blockers', 'Blocker', 'Note', 'Warning', 'Risk', 'Limitation'}


def _apply_bold(text: str) -> str:
    """Convert **text** to <strong>text</strong> with colored labels.

    Suffix convention (from scorer.py):
      **Label — Blocker:** → label in red   (#e05252)
      **Label — Risk:**    → label in orange (#f59e0b)
    Fallback: first-word matches against _WARNING_LABELS → red.
    """
    def _replace(m):
        label = m.group(1)
        colon_pos = label.find(':')
        label_text = label[:colon_pos].strip() if colon_pos != -1 else label.strip()
        rest = label[colon_pos:] if colon_pos != -1 else ''

        # Suffix convention — check before first-word fallback
        if label_text.endswith('— Blocker') or label_text.endswith('\u2014 Blocker'):
            return f'<strong><span style="color:#e05252;">{label_text}</span>{rest}</strong>'
        if label_text.endswith('— Risk') or label_text.endswith('\u2014 Risk'):
            return f'<strong><span style="color:#f59e0b;">{label_text}</span>{rest}</strong>'

        # Legacy first-word check
        first_word = label_text.split()[0] if label_text else label_text
        if first_word in _WARNING_LABELS:
            return f'<strong><span style="color:#e05252;">{label_text}</span>{rest}</strong>'

        return f'<strong>{label}</strong>'
    return _re.sub(r'\*\*(.+?)\*\*', _replace, text)


@app.template_filter('bold_labels')
def bold_labels(text):
    """Convert **text** markdown to <strong>text</strong> HTML.
    Labels matching warning patterns render in red."""
    return Markup(_apply_bold(str(text)))


@app.template_filter('linkify')
def linkify(text):
    """Convert [label](url) markdown links to <a> HTML, then apply bold with warning colors."""
    result = str(text)
    result = _re.sub(
        r'\[([^\]]+)\]\((https?://[^\)]+)\)',
        r'<a href="\2" target="_blank" class="rec-link">\1</a>',
        result,
    )
    result = _apply_bold(result)
    return Markup(result)


@app.template_filter('link_product_name')
def link_product_name(text, name, url):
    """Wrap the first occurrence of the product name in a description with a link."""
    if not url or not name:
        return Markup(str(text))
    safe_text = str(_escape(text))
    safe_url = str(_escape(url))
    safe_name = str(_escape(name))
    linked = safe_text.replace(safe_name, f'<a href="{safe_url}" target="_blank" class="product-name-link">{safe_name}</a>', 1)
    return Markup(linked)


# ---------------------------------------------------------------------------
# In-memory progress store for SSE streaming
# ---------------------------------------------------------------------------

_progress: dict[str, list[str]] = {}
_progress_lock = threading.Lock()
_PROGRESS_MAX_JOBS = 50
_PROGRESS_EVICT_COUNT = 10
_cancelled_jobs: set[str] = set()


def _push(job_id: str, msg: str):
    with _progress_lock:
        if job_id not in _progress:
            if len(_progress) >= _PROGRESS_MAX_JOBS:
                for old_key in list(_progress.keys())[:_PROGRESS_EVICT_COUNT]:
                    del _progress[old_key]
            _progress[job_id] = []
        _progress[job_id].append(msg)


def _sse_stream(job_id: str, poll_interval: float = 0.3):
    """Generic SSE generator: yields messages for job_id until done/error or timeout."""
    last = 0
    deadline = time.time() + _SSE_TIMEOUT
    while time.time() < deadline:
        with _progress_lock:
            msgs = _progress.get(job_id, [])
        for msg in msgs[last:]:
            last += 1
            yield f"data: {msg}\n\n"
            if msg.startswith("done:") or msg.startswith("error:"):
                return
        time.sleep(poll_interval)
    yield "data: error:Timed out — the operation took too long. Please try again.\n\n"


# ---------------------------------------------------------------------------
# Composite score helper (single source of truth)
# ---------------------------------------------------------------------------

_CHANNEL_ORGS = {"training_organization", "systems_integrator", "technology_distributor", "professional_services", "academic_institution"}


def _compute_composite(top_lab: int, pr_score: int, org_type: str) -> tuple[int, bool, str]:
    """Return (composite_score, is_gated, weights_label)."""
    if org_type in _CHANNEL_ORGS:
        raw = round(top_lab * 0.35 + pr_score * 0.65)
        gate_threshold, gate_cap = 20, 30
        weights = "35% Product / 65% Partnership"
    else:
        raw = round(top_lab * 0.65 + pr_score * 0.35)
        gate_threshold, gate_cap = 30, 25
        weights = "65% Product / 35% Partnership"

    if top_lab < gate_threshold:
        return min(raw, gate_cap), True, weights
    return min(100, raw), False, weights


# ---------------------------------------------------------------------------
# Platform landing page
# ---------------------------------------------------------------------------

@app.route("/")
def platform_home():
    recent = list_analyses()
    return render_template("home.html", recent=recent, platform_mode=True)


# ---------------------------------------------------------------------------
# Inspector Blueprint
# ---------------------------------------------------------------------------

inspector = Blueprint("inspector", __name__, url_prefix="/inspector")


@inspector.route("/")
@inspector.route("")
def home():
    recent = list_analyses()
    return render_template("home.html", recent=recent)


# Phase 1: Discovery

@inspector.route("/discover", methods=["POST"])
def discover():
    company_name = request.form.get("company_name", "").strip()
    product_name = request.form.get("product_name", "").strip()

    if not company_name and not product_name:
        return redirect(url_for("inspector.home"))

    if company_name and product_name:
        known_products = [product_name]
        search_label = f"{company_name} {product_name}"
    elif product_name and not company_name:
        company_name, known_products = resolve_company_from_product(product_name)
        search_label = product_name
    else:
        known_products = None
        search_label = company_name

    discovery_id = str(uuid.uuid4())[:8]

    def run_discovery():
        try:
            _push(discovery_id, "status:Searching for products...")
            findings = discover_products(company_name, known_products)
            _push(discovery_id, "status:Analyzing product portfolio...")
            discovery = discover_products_with_claude(findings)
            discovery["discovery_id"] = discovery_id
            discovery["known_products"] = known_products or []
            for key in ("training_programs", "atp_signals", "training_catalog",
                        "partner_ecosystem", "partner_portal", "cs_signals",
                        "lms_signals", "org_contacts", "page_contents"):
                discovery[key] = findings.get(key, [])
            save_discovery(discovery_id, discovery)
            _push(discovery_id, f"done:{discovery_id}")
        except Exception as e:
            traceback.print_exc()
            _push(discovery_id, f"error:{e}")

    threading.Thread(target=run_discovery, daemon=True).start()
    return render_template("discovering.html", discovery_id=discovery_id, search_label=search_label)


@inspector.route("/discover/progress/<discovery_id>")
def discover_progress(discovery_id: str):
    return Response(stream_with_context(_sse_stream(discovery_id)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@inspector.route("/select/<discovery_id>")
def select_products(discovery_id: str):
    discovery = load_discovery(discovery_id)
    if not discovery:
        return redirect(url_for("inspector.home"))
    for i, p in enumerate(discovery.get("products", []), start=1):
        if not p.get("priority"):
            p["priority"] = i
    existing_analysis = find_analysis_by_discovery_id(discovery_id)
    return render_template("select.html", discovery=discovery, existing_analysis=existing_analysis)


# Phase 2: Scoring

@inspector.route("/score", methods=["POST"])
def score():
    discovery_id = request.form.get("discovery_id", "")
    discovery = load_discovery(discovery_id)
    if not discovery:
        return redirect(url_for("inspector.home"))

    selected_names = request.form.getlist("products")
    all_products = discovery.get("products", [])
    selected_products = [p for p in all_products if p["name"] in selected_names]

    if not selected_products:
        return redirect(url_for("inspector.select_products", discovery_id=discovery_id))

    job_id = str(uuid.uuid4())[:8]
    company_name = discovery.get("company_name", "")

    def run_scoring():
        try:
            # Check if research is fully cached for all selected products
            cache = discovery.get("_research_cache", {})
            cached_names = set(cache.get("researched_products", []))
            selected_names = {p["name"] for p in selected_products}

            if selected_names <= cached_names:
                # All selected products already researched — skip web searches
                _push(job_id, "status:Loading cached research data")
                research = {
                    "company_name": company_name,
                    "selected_products": selected_products,
                    "search_results": cache.get("search_results", {}),
                    "page_contents": cache.get("page_contents", {}),
                }
            else:
                research_messages = [
                    "status:Scouring technical docs, APIs & deployment guides",
                    "status:Finding training (ATP and gray) & certification programs",
                    "status:Researching product focus areas, AI features & target personas",
                    "status:Analyzing channel partners, enablement resources, roadshows & hands-on events",
                    "status:Identifying training and enablement organizational structures & contacts",
                ]
                def push_research_messages():
                    for i, msg in enumerate(research_messages):
                        _push(job_id, msg)
                        if i < len(research_messages) - 1:
                            threading.Event().wait(5)

                msg_thread = threading.Thread(target=push_research_messages, daemon=True)
                msg_thread.start()

                research = research_products(company_name, selected_products)
                msg_thread.join()

                # Save research to discovery cache for future re-runs
                merged_search = {**cache.get("search_results", {}), **research["search_results"]}
                merged_pages = {**cache.get("page_contents", {}), **research["page_contents"]}
                updated_discovery = {**discovery, "_research_cache": {
                    "researched_products": list(cached_names | selected_names),
                    "search_results": merged_search,
                    "page_contents": merged_pages,
                }}
                save_discovery(discovery_id, updated_discovery)

            research["discovery_data"] = discovery

            _push(job_id, "status:Scoring with Claude AI")
            analysis = score_selected_products(research)
            analysis.discovery_id = discovery_id
            analysis.total_products_discovered = len(discovery.get("products", []))
            analysis_id = save_analysis(analysis)
            _push(job_id, "status:Generating report")
            _push(job_id, f"done:{analysis_id}")
        except Exception as e:
            traceback.print_exc()
            _push(job_id, f"error:{e}")

    threading.Thread(target=run_scoring, daemon=True).start()
    return render_template("scoring.html", job_id=job_id, company_name=company_name,
                           product_count=len(selected_products))


@inspector.route("/score/progress/<job_id>")
def score_progress(job_id: str):
    return Response(stream_with_context(_sse_stream(job_id)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# Results

@inspector.route("/results/<analysis_id>")
def results(analysis_id: str):
    data = load_analysis(analysis_id)
    if not data:
        return "Analysis not found", 404

    products = data.get("products") or []
    for p in products:
        scores = p.get("labability_score") or {}
        tech = scores.get("technical_orchestrability", {}).get("score", 0) if scores else 0
        other = sum(
            s.get("score", 0)
            for k, s in scores.items()
            if k != "technical_orchestrability" and isinstance(s, dict)
        ) if scores else 0
        p["_total_score"] = compute_labability_total(tech, other, p.get("skillable_path", ""))
    products.sort(key=lambda p: p.get("_total_score", 0), reverse=True)
    data["products"] = products

    pr = data.get("partnership_readiness") or {}
    pr_raw = sum(s.get("score", 0) for s in pr.values() if isinstance(s, dict))
    data["_pr_total"] = min(100, round(pr_raw / 1.17))

    org_type = data.get("organization_type", "software_company")
    top_lab = products[0].get("_total_score", 0) if products else 0
    composite, gated, weights = _compute_composite(top_lab, data["_pr_total"], org_type)
    data["_composite_score"] = composite
    data["_composite_gated"] = gated
    data["_composite_weights"] = weights or "65% Product / 35% Partnership"

    return render_template("results.html", data=data)


# CSV export

@inspector.route("/export/<analysis_id>")
def export(analysis_id: str):
    data = load_analysis(analysis_id)
    if not data:
        return "Analysis not found", 404

    products = data.get("products") or []
    for p in products:
        scores = p.get("labability_score") or {}
        tech = scores.get("technical_orchestrability", {}).get("score", 0) if scores else 0
        other = sum(s.get("score", 0) for k, s in scores.items()
                    if k != "technical_orchestrability" and isinstance(s, dict))
        p["_total_score"] = compute_labability_total(tech, other, p.get("skillable_path", ""))

    pr = data.get("partnership_readiness") or {}
    pr_raw = sum(s.get("score", 0) for s in pr.values() if isinstance(s, dict))
    pr_total = min(100, round(pr_raw / 1.17))

    output = io.StringIO()
    fieldnames = ["company_name", "product_name", "labability_score", "partnership_score",
                  "composite_score", "skillable_path", "org_type"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    _csv_defaults = {f: "" for f in fieldnames}
    writer.writeheader()
    org_type = data.get("organization_type", "software_company")
    for p in products:
        lab = p.get("_total_score", 0)
        composite, _, _ = _compute_composite(lab, pr_total, org_type)
        writer.writerow({**_csv_defaults,
                         "company_name": data.get("company_name", ""),
                         "product_name": p.get("name", ""),
                         "labability_score": lab,
                         "partnership_score": pr_total,
                         "composite_score": composite,
                         "skillable_path": p.get("skillable_path", ""),
                         "org_type": org_type})

    output.seek(0)
    safe_name = data.get("company_name", "analysis").replace(" ", "-").lower()
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=inspector-{safe_name}-{analysis_id}.csv"},
    )


# API (batch mode)

@inspector.route("/api/analyze", methods=["POST"])
def api_analyze():
    body = request.get_json()
    company_name = body.get("company_name", "").strip()
    known_products = body.get("known_products")
    selected_product_names = body.get("products")

    if not company_name:
        return jsonify({"error": "company_name is required"}), 400

    try:
        findings = discover_products(company_name, known_products)
        discovery = discover_products_with_claude(findings)
        for key in ("training_programs", "atp_signals", "training_catalog",
                    "partner_ecosystem", "partner_portal", "cs_signals",
                    "lms_signals", "org_contacts", "page_contents"):
            discovery[key] = findings.get(key, [])

        if selected_product_names:
            selected = [p for p in discovery["products"] if p["name"] in selected_product_names]
        else:
            selected = discovery["products"]

        research = research_products(company_name, selected)
        research["discovery_data"] = discovery
        analysis = score_selected_products(research)
        analysis_id = save_analysis(analysis)
        return jsonify(load_analysis(analysis_id))
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Prospector Blueprint
# ---------------------------------------------------------------------------

prospector = Blueprint("prospector", __name__, url_prefix="/prospector")


def _derive_skillable_path(lab_score: int) -> str:
    """Map a lab score to the three Prospector path labels."""
    if lab_score >= 50:
        return "Labable"
    if lab_score >= 20:
        return "Simulations"
    return "Do Not Pursue"


@prospector.route("/")
@prospector.route("")
def prospector_home():
    poor_fit = load_poor_fit_companies()
    return render_template("prospector.html", poor_fit_count=len(poor_fit))


@prospector.route("/run", methods=["POST"])
def prospector_run():
    company_names = []

    # Mode 1: CSV file upload — read first column
    uploaded = request.files.get("csv_file")
    if uploaded and uploaded.filename:
        stream = io.StringIO(uploaded.stream.read().decode("utf-8-sig"))
        reader = csv.reader(stream)
        for row in reader:
            if row and row[0].strip() and row[0].strip().lower() != "company":
                company_names.append(row[0].strip())

    # Mode 2: pasted text (fallback)
    if not company_names:
        raw = request.form.get("companies", "")
        company_names = [n.strip() for n in raw.replace(",", "\n").splitlines() if n.strip()]

    if not company_names:
        return redirect(url_for("prospector.prospector_home"))

    # Filter out previously flagged poor fits
    poor_fit = load_poor_fit_companies()
    filtered = [n for n in company_names if n.lower() not in poor_fit]
    skipped = len(company_names) - len(filtered)
    company_names = filtered

    if not company_names:
        return redirect(url_for("prospector.prospector_home"))

    job_id = str(uuid.uuid4())[:8]
    results = {}
    semaphore = threading.Semaphore(3)

    def run_batch():
        def analyze_one(name):
            if job_id in _cancelled_jobs:
                return
            with semaphore:
                if job_id in _cancelled_jobs:
                    return
                _push(job_id, f"status:Analyzing {name}...")
                row = _quick_analyze_company(name)
                if row:
                    row["skillable_path"] = _derive_skillable_path(row.get("lab_score", 0))
                    row["flagged_poor_fit"] = False
                results[name] = row or {"company_name": name, "error": "Analysis failed"}
                _push(job_id, f"progress:{name}")

        threads = [threading.Thread(target=analyze_one, args=(n,), daemon=True) for n in company_names]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        cancelled = job_id in _cancelled_jobs
        _cancelled_jobs.discard(job_id)
        save_prospector_run(job_id, {
            "job_id": job_id,
            "companies": company_names,
            "skipped_poor_fit": skipped,
            "results": results,
            "status": "cancelled" if cancelled else "done",
        })
        _push(job_id, f"done:{job_id}")

    threading.Thread(target=run_batch, daemon=True).start()
    return render_template("prospector_running.html", job_id=job_id,
                           company_names=company_names,
                           company_count=len(company_names))


@prospector.route("/cancel/<job_id>", methods=["POST"])
def prospector_cancel(job_id: str):
    _cancelled_jobs.add(job_id)
    _push(job_id, f"done:{job_id}")
    return jsonify({"cancelled": True})


@prospector.route("/status/<job_id>")
def prospector_status(job_id: str):
    return Response(stream_with_context(_sse_stream(job_id, poll_interval=0.5)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@prospector.route("/results/<job_id>")
def prospector_results(job_id: str):
    job = load_prospector_run(job_id)
    if not job:
        return redirect(url_for("prospector.prospector_home"))
    results = job.get("results", {})
    rows = [r for r in results.values() if r and "error" not in r]
    rows.sort(key=lambda r: r.get("composite_score", 0), reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    errors = [r for r in results.values() if r and "error" in r]
    return render_template("prospector_results.html", rows=rows, errors=errors,
                           job_id=job_id,
                           company_count=len(job.get("companies", [])),
                           skipped_poor_fit=job.get("skipped_poor_fit", 0))


@prospector.route("/flag", methods=["POST"])
def prospector_flag():
    data = request.get_json() or {}
    company_name = data.get("company_name", "").strip()
    job_id = data.get("job_id", "").strip()
    reason = data.get("reason", "").strip()
    if not company_name:
        return jsonify({"error": "company_name required"}), 400

    from datetime import datetime
    append_poor_fit_feedback({
        "company_name": company_name,
        "job_id": job_id,
        "reason": reason,
        "flagged_at": datetime.now().isoformat(),
    })

    # Mark row as flagged in the stored job
    job = load_prospector_run(job_id)
    if job and company_name in job.get("results", {}):
        job["results"][company_name]["flagged_poor_fit"] = True
        save_prospector_run(job_id, job)

    return jsonify({"success": True})


@prospector.route("/export/<job_id>")
def prospector_export(job_id: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    job = load_prospector_run(job_id)
    if not job:
        return "Job not found", 404

    results = job.get("results", {})
    rows = sorted([r for r in results.values() if r and "error" not in r],
                  key=lambda r: r.get("composite_score", 0), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospector Results"

    # Color palette
    green_dark  = "FF0D1A14"
    green_hi    = "FF24ED9B"
    green_mid   = "FF6b9e88"
    amber       = "FFF59E0B"
    red_soft    = "FFF87171"
    white       = "FFE8F5F0"
    muted       = "FF3d6655"
    bg_header   = "FF06100C"

    header_font  = Font(bold=True, color=green_hi, size=9)
    default_font = Font(color=white, size=9)
    muted_font   = Font(color=muted, size=8)
    thin_side    = Side(style="thin", color="1E3329")
    thin_border  = Border(bottom=thin_side)

    headers = ["#", "Company", "Top Product", "Skillable Path",
               "Lab Score", "Partnership", "Composite",
               "Highly Labable Products", "Likely Labable Products", "Not Labable Products",
               "ATP Program", "Channel Program", "On-Demand Library", "Cert Program",
               "Existing Lab Partner",
               "Top Contact", "Title", "LinkedIn", "Full Analysis", "Notes"]
    col_widths = [4, 26, 28, 16, 11, 11, 11, 10, 10, 10, 26, 26, 14, 12, 18, 22, 24, 12, 14, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=bg_header)
        cell.alignment = Alignment(horizontal="center" if col in (1,5,6,7,10,11) else "left",
                                   vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    path_colors = {"Labable": green_hi, "Simulations": amber, "Do Not Pursue": red_soft}

    for i, row in enumerate(rows, 2):
        lab   = row.get("lab_score", 0)
        prtn  = row.get("partnership_score", 0)
        comp  = row.get("composite_score", 0)
        path  = row.get("skillable_path", "")
        aid   = row.get("analysis_id", "")

        def score_color(s):
            return green_hi if s >= 70 else (amber if s >= 40 else red_soft)

        values = [
            row.get("rank", i - 1),
            row.get("company_name", ""),
            row.get("top_product", ""),
            path,
            lab, prtn, comp,
            row.get("total_highly_labable", ""),
            row.get("total_likely_labable", ""),
            row.get("total_not_labable", ""),
            row.get("atp_program", ""),
            row.get("channel_program", ""),
            row.get("ondemand_library", ""),
            row.get("cert_program", ""),
            row.get("existing_lab_partner", ""),
            row.get("top_contact_name", ""),
            row.get("top_contact_title", ""),
            row.get("top_contact_linkedin", ""),
            f"/inspector/results/{aid}" if aid else "",
            "",  # Notes — blank for user to fill
        ]
        # Center columns: #(1), scores(5,6,7), product counts(8,9,10), ondemand(13), cert(14)
        center_cols = {1, 5, 6, 7, 8, 9, 10, 13, 14}
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor="FF0D1A14")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col in center_cols else "left")
            if col in (5, 6, 7):
                cell.font = Font(bold=True, color=score_color(val if isinstance(val, int) else 0), size=9)
            elif col == 4:
                cell.font = Font(bold=True, color=path_colors.get(path, white), size=9)
            elif col == 1:
                cell.font = Font(color=muted, size=9, bold=True)
            elif col in (8, 9, 10):
                cell.font = Font(color=green_mid, size=9)
            else:
                cell.font = default_font
        ws.row_dimensions[i].height = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=prospector-{job_id}.xlsx"},
    )


# ---------------------------------------------------------------------------
# Designer Blueprint (stub — Phase 3)
# ---------------------------------------------------------------------------

designer = Blueprint("designer", __name__, url_prefix="/designer")


@designer.route("/")
@designer.route("")
def designer_home():
    return render_template("designer.html")


# ---------------------------------------------------------------------------
# Prospector batch (currently served under /inspector/marketing — kept for compat)
# Will move fully to /prospector in Phase 4
# ---------------------------------------------------------------------------

_LABABLE_ORDER = {"highly_likely": 0, "likely": 1, "less_likely": 2, "not_likely": 3}


def _quick_analyze_company(company_name: str) -> dict | None:
    """Run full discovery + score for a single company (top 1 product)."""
    try:
        findings = discover_products(company_name)
        discovery = discover_products_with_claude(findings)
        for key in ("training_programs", "atp_signals", "training_catalog",
                    "partner_ecosystem", "partner_portal", "cs_signals",
                    "lms_signals", "org_contacts", "page_contents"):
            discovery[key] = findings.get(key, [])

        all_products = discovery.get("products", [])
        labable = [p for p in all_products if p.get("likely_labable") != "not_likely"]
        if not labable:
            labable = all_products
        labable.sort(key=lambda p: _LABABLE_ORDER.get(p.get("likely_labable", "not_likely"), 4))
        selected = labable[:1]
        if not selected:
            return None

        research = research_products(company_name, selected)
        research["discovery_data"] = discovery
        analysis = score_selected_products(research)
        analysis_id = save_analysis(analysis)

        data = load_analysis(analysis_id)
        products = data.get("products") or []
        for p in products:
            scores = p.get("labability_score") or {}
            tech = scores.get("technical_orchestrability", {}).get("score", 0)
            other = sum(s.get("score", 0) for k, s in scores.items()
                        if k != "technical_orchestrability" and isinstance(s, dict))
            p["_total_score"] = compute_labability_total(tech, other, p.get("skillable_path", ""))
        products.sort(key=lambda p: p.get("_total_score", 0), reverse=True)

        pr = data.get("partnership_readiness") or {}
        pr_raw = sum(s.get("score", 0) for s in pr.values() if isinstance(s, dict))
        pr_total = min(100, round(pr_raw / 1.17))

        top_product = products[0] if products else {}
        lab_score = top_product.get("_total_score", 0)
        org_type = data.get("organization_type", "software_company")
        composite, _, _ = _compute_composite(lab_score, pr_total, org_type)

        contacts = top_product.get("contacts") or []
        dm = next((c for c in contacts if c.get("role_type") == "decision_maker"), None) or (contacts[0] if contacts else None)

        # Product labability counts from discovery phase
        all_disc = discovery.get("products", [])
        total_highly  = sum(1 for p in all_disc if p.get("likely_labable") == "highly_likely")
        total_likely  = sum(1 for p in all_disc if p.get("likely_labable") == "likely")
        total_not     = sum(1 for p in all_disc if p.get("likely_labable") in ("less_likely", "not_likely"))

        # Partnership signals from discovery Claude output
        ps = discovery.get("partnership_signals", {})

        def _fmt_ondemand(val):
            if val is None: return ""
            if val == -1: return "Yes"
            return str(val) if val > 0 else ""

        def _fmt_cert(val):
            if val is None: return ""
            return str(val)

        return {
            "company_name": data.get("company_name", company_name),
            "company_url": data.get("company_url", ""),
            "top_product": top_product.get("name", ""),
            "lab_score": lab_score,
            "partnership_score": pr_total,
            "composite_score": composite,
            "top_contact_name": dm.get("name", "") if dm else "",
            "top_contact_title": dm.get("title", "") if dm else "",
            "top_contact_linkedin": dm.get("linkedin_url", "") if dm else "",
            "analysis_id": analysis_id,
            "total_highly_labable": total_highly,
            "total_likely_labable": total_likely,
            "total_not_labable": total_not,
            "atp_program": ps.get("atp_program") or "",
            "channel_program": ps.get("channel_program") or "",
            "ondemand_library": _fmt_ondemand(ps.get("ondemand_library")),
            "cert_program": _fmt_cert(ps.get("cert_program")),
            "existing_lab_partner": ps.get("existing_lab_partner") or "",
        }
    except Exception:
        traceback.print_exc()
        return None


@inspector.route("/marketing")
def marketing():
    return redirect(url_for("prospector.prospector_home"), 301)


@inspector.route("/marketing/run", methods=["POST"])
def marketing_run():
    raw = request.form.get("companies", "")
    company_names = [n.strip() for n in raw.replace(",", "\n").splitlines() if n.strip()]
    if not company_names:
        return redirect(url_for("inspector.marketing"))

    job_id = str(uuid.uuid4())[:8]
    save_batch_job(job_id, {"job_id": job_id, "companies": company_names, "results": {}, "status": "running"})

    def run_batch():
        results = {}
        semaphore = threading.Semaphore(3)

        def analyze_one(name):
            with semaphore:
                _push(job_id, f"status:Analyzing {name}...")
                row = _quick_analyze_company(name)
                results[name] = row or {"company_name": name, "error": "Analysis failed"}
                _push(job_id, f"progress:{name}")

        threads = [threading.Thread(target=analyze_one, args=(n,), daemon=True) for n in company_names]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        save_batch_job(job_id, {"job_id": job_id, "companies": company_names, "results": results, "status": "done"})
        _push(job_id, f"done:{job_id}")

    threading.Thread(target=run_batch, daemon=True).start()
    return render_template("marketing_running.html", job_id=job_id,
                           company_count=len(company_names))


@inspector.route("/marketing/status/<job_id>")
def marketing_status(job_id: str):
    return Response(stream_with_context(_sse_stream(job_id, poll_interval=0.5)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@inspector.route("/marketing/results/<job_id>")
def marketing_results(job_id: str):
    job = load_batch_job(job_id)
    if not job:
        return redirect(url_for("inspector.marketing"))
    results = job.get("results", {})
    rows = [r for r in results.values() if r and "error" not in r]
    rows.sort(key=lambda r: r.get("composite_score", 0), reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    errors = [r for r in results.values() if r and "error" in r]
    return render_template("marketing_results.html", rows=rows, errors=errors,
                           job_id=job_id, company_count=len(job.get("companies", [])))


@inspector.route("/marketing/export/<job_id>")
def marketing_export(job_id: str):
    job = load_batch_job(job_id)
    if not job:
        return "Job not found", 404
    results = job.get("results", {})
    rows = sorted([r for r in results.values() if r and "error" not in r],
                  key=lambda r: r.get("composite_score", 0), reverse=True)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "rank", "company_name", "company_url", "top_product",
        "lab_score", "partnership_score", "composite_score",
        "top_contact_name", "top_contact_title", "top_contact_linkedin", "analysis_id",
    ])
    _csv_defaults = {f: "" for f in writer.fieldnames}
    writer.writeheader()
    for i, r in enumerate(rows):
        writer.writerow({**_csv_defaults, **r, "rank": i + 1})

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=labability-batch-{job_id}.csv"},
    )


# Register all blueprints after all routes are defined
app.register_blueprint(inspector)
app.register_blueprint(prospector)
app.register_blueprint(designer)


if __name__ == "__main__":
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("\n*** WARNING: ANTHROPIC_API_KEY not set. Set it in .env or environment. ***\n")
    app.run(debug=True, port=5000)
